from openpyxl import *
from datetime import date, timedelta, datetime
import pandas as pd
from os import path, startfile
from setup_data import setup_data
import urllib
def generate():
    print("Generating...")
    start_date=date(2020,1,22)
    end_date=str(datetime.today()).split()[0]
    tdate1=date(int(end_date[:4]),int(end_date[5:7]),int(end_date[8:]))
    delta=timedelta(days=1)
    tdate=(tdate1-delta).strftime("%m-%d-%Y")
    BASE_URL="https://raw.githubusercontent.com/CSSEGISandData/COVID-19/master/csse_covid_19_data/csse_covid_19_daily_reports/"
    countries=['Afghanistan', 'Albania', 'Algeria', 'Andorra', 'Angola', 'Antigua and Barbuda', 'Argentina', 'Armenia', 'Australia', 'Austria', 'Azerbaijan', 'Bahamas', 'Bahrain', 'Bangladesh', 'Barbados', 'Belarus', 'Belgium', 'Belize', 'Benin', 'Bhutan', 'Bolivia', 'Bosnia and Herzegovina', 'Botswana', 'Brazil', 'Brunei', 'Bulgaria', 'Burkina Faso', 'Burma', 'Burundi', 'Cabo Verde', 'Cambodia', 'Cameroon', 'Canada', 'Central African Republic', 'Chad', 'Chile', 'China', 'Colombia', 'Comoros', 'Congo (Brazzaville)', 'Congo (Kinshasa)', 'Costa Rica', "Cote d'Ivoire", 'Croatia', 'Cuba', 'Cyprus', 'Czechia', 'Denmark', 'Diamond Princess', 'Djibouti', 'Dominica', 'Dominican Republic', 'Ecuador', 'Egypt', 'El Salvador', 'Equatorial Guinea', 'Eritrea', 'Estonia', 'Eswatini', 'Ethiopia', 'Fiji', 'Finland', 'France', 'Gabon', 'Gambia', 'Georgia', 'Germany', 'Ghana', 'Greece', 'Grenada', 'Guatemala', 'Guinea', 'Guinea-Bissau', 'Guyana', 'Haiti', 'Holy See', 'Honduras', 'Hungary', 'Iceland', 'India', 'Indonesia', 'Iran', 'Iraq', 'Ireland', 'Israel', 'Italy', 'Jamaica', 'Japan', 'Jordan', 'Kazakhstan', 'Kenya', 'Kiribati', 'Korea, South', 'Kosovo', 'Kuwait', 'Kyrgyzstan', 'Laos', 'Latvia', 'Lebanon', 'Lesotho', 'Liberia', 'Libya', 'Liechtenstein', 'Lithuania', 'Luxembourg', 'MS Zaandam', 'Madagascar', 'Malawi', 'Malaysia', 'Maldives', 'Mali', 'Malta', 'Marshall Islands', 'Mauritania', 'Mauritius', 'Mexico', 'Micronesia', 'Moldova', 'Monaco', 'Mongolia', 'Montenegro', 'Morocco', 'Mozambique', 'Namibia', 'Nauru', 'Nepal', 'Netherlands', 'New Zealand', 'Nicaragua', 'Niger', 'Nigeria', 'North Korea', 'North Macedonia', 'Norway', 'Oman', 'Pakistan', 'Palau', 'Palestine State', 'Panama', 'Papua New Guinea', 'Paraguay', 'Peru', 'Philippines', 'Poland', 'Portugal', 'Qatar', 'Romania', 'Russia', 'Rwanda', 'Saint Kitts and Nevis', 'Saint Lucia', 'Saint Vincent and the Grenadines', 'Samoa', 'San Marino', 'Sao Tome and Principe', 'Saudi Arabia', 'Senegal', 'Serbia', 'Seychelles', 'Sierra Leone', 'Singapore', 'Slovakia', 'Slovenia', 'Solomon Islands', 'Somalia', 'South Africa', 'South Sudan', 'Spain', 'Sri Lanka', 'Sudan', 'Suriname', 'Sweden', 'Switzerland', 'Syria', 'Taiwan', 'Tajikistan', 'Tanzania', 'Thailand', 'Timor-Leste', 'Togo', 'Tonga', 'Trinidad and Tobago', 'Tunisia', 'Turkey', 'Turkmenistan', 'Tuvalu', 'US', 'Uganda', 'Ukraine', 'United Arab Emirates', 'United Kingdom', 'Uruguay', 'Uzbekistan', 'Vanuatu', 'Venezuela', 'Vietnam', 'West Bank and Gaza', 'Western Sahara', 'Yemen', 'Zambia', 'Zimbabwe']
    l=utils.cell.get_column_interval(2,240)
    names=["Active","Recovered","Confirmed","Deaths"]
    book=Workbook()
    for t in range(4):
        start_date=date(2020,1,22)
        end_date=str(datetime.today()).split()[0]
        tdate1=date(int(end_date[:4]),int(end_date[5:7]),int(end_date[8:]))
        delta=timedelta(days=1)
        tdate=(tdate1-delta).strftime("%m-%d-%Y")
        if t==0:
            wb=book.active
        else:
            wb=book.create_sheet('s')
        print(names[t])
        wb.title=names[t]
        i=2
        for a in range(len(countries)):
            wb[l[a]+'1']=countries[a]
            #print(countries[a])
        while start_date<=tdate1:
            dat=start_date.strftime("%m-%d-%Y")
            try:
                data=setup_data(pd.read_csv(BASE_URL+dat+".csv"))
            except urllib.error.HTTPError:
                break
            wb['A'+str(i)]=dat
            for a in range(len(countries)):
                x=list(data.index)
                y=countries[a]
                if y in x:
                    wb[l[a]+str(i)]=int(data.loc[y][names[t]])
                elif y not in x:
                    f=False
                    for b in range(len(x)):
                        if y in x[b]:
                            f=True
                            y=x[b]
                            break
                    if f:
                        wb[l[a]+str(i)]=int(data.loc[y][names[t]])
                    else:
                        wb[l[a]+str(i)]=0
                else:
                    wb[l[a]+'2']=0
            start_date+=delta
            print(dat)
            i+=1
        wb['A1']="Date"
    g=open("STATUS.txt",'w')
    g.write(dat+' '+str(i))
    g.close()
    book.save(filename="data.xlsx")
def update():
    BASE_URL="https://raw.githubusercontent.com/CSSEGISandData/COVID-19/master/csse_covid_19_data/csse_covid_19_daily_reports/"
    countries=['Afghanistan', 'Albania', 'Algeria', 'Andorra', 'Angola', 'Antigua and Barbuda', 'Argentina', 'Armenia', 'Australia', 'Austria', 'Azerbaijan', 'Bahamas', 'Bahrain', 'Bangladesh', 'Barbados', 'Belarus', 'Belgium', 'Belize', 'Benin', 'Bhutan', 'Bolivia', 'Bosnia and Herzegovina', 'Botswana', 'Brazil', 'Brunei', 'Bulgaria', 'Burkina Faso', 'Burma', 'Burundi', 'Cabo Verde', 'Cambodia', 'Cameroon', 'Canada', 'Central African Republic', 'Chad', 'Chile', 'China', 'Colombia', 'Comoros', 'Congo (Brazzaville)', 'Congo (Kinshasa)', 'Costa Rica', "Cote d'Ivoire", 'Croatia', 'Cuba', 'Cyprus', 'Czechia', 'Denmark', 'Diamond Princess', 'Djibouti', 'Dominica', 'Dominican Republic', 'Ecuador', 'Egypt', 'El Salvador', 'Equatorial Guinea', 'Eritrea', 'Estonia', 'Eswatini', 'Ethiopia', 'Fiji', 'Finland', 'France', 'Gabon', 'Gambia', 'Georgia', 'Germany', 'Ghana', 'Greece', 'Grenada', 'Guatemala', 'Guinea', 'Guinea-Bissau', 'Guyana', 'Haiti', 'Holy See', 'Honduras', 'Hungary', 'Iceland', 'India', 'Indonesia', 'Iran', 'Iraq', 'Ireland', 'Israel', 'Italy', 'Jamaica', 'Japan', 'Jordan', 'Kazakhstan', 'Kenya', 'Kiribati', 'Korea, South', 'Kosovo', 'Kuwait', 'Kyrgyzstan', 'Laos', 'Latvia', 'Lebanon', 'Lesotho', 'Liberia', 'Libya', 'Liechtenstein', 'Lithuania', 'Luxembourg', 'MS Zaandam', 'Madagascar', 'Malawi', 'Malaysia', 'Maldives', 'Mali', 'Malta', 'Marshall Islands', 'Mauritania', 'Mauritius', 'Mexico', 'Micronesia', 'Moldova', 'Monaco', 'Mongolia', 'Montenegro', 'Morocco', 'Mozambique', 'Namibia', 'Nauru', 'Nepal', 'Netherlands', 'New Zealand', 'Nicaragua', 'Niger', 'Nigeria', 'North Korea', 'North Macedonia', 'Norway', 'Oman', 'Pakistan', 'Palau', 'Palestine State', 'Panama', 'Papua New Guinea', 'Paraguay', 'Peru', 'Philippines', 'Poland', 'Portugal', 'Qatar', 'Romania', 'Russia', 'Rwanda', 'Saint Kitts and Nevis', 'Saint Lucia', 'Saint Vincent and the Grenadines', 'Samoa', 'San Marino', 'Sao Tome and Principe', 'Saudi Arabia', 'Senegal', 'Serbia', 'Seychelles', 'Sierra Leone', 'Singapore', 'Slovakia', 'Slovenia', 'Solomon Islands', 'Somalia', 'South Africa', 'South Sudan', 'Spain', 'Sri Lanka', 'Sudan', 'Suriname', 'Sweden', 'Switzerland', 'Syria', 'Taiwan', 'Tajikistan', 'Tanzania', 'Thailand', 'Timor-Leste', 'Togo', 'Tonga', 'Trinidad and Tobago', 'Tunisia', 'Turkey', 'Turkmenistan', 'Tuvalu', 'US', 'Uganda', 'Ukraine', 'United Arab Emirates', 'United Kingdom', 'Uruguay', 'Uzbekistan', 'Vanuatu', 'Venezuela', 'Vietnam', 'West Bank and Gaza', 'Western Sahara', 'Yemen', 'Zambia', 'Zimbabwe']
    l=utils.cell.get_column_interval(2,240)
    print("Updating...")
    book=load_workbook("data.xlsx")
    wb=book.get_sheet_by_name("COVID-19 Information By Country")
    f=open("STATUS.txt",'r')
    inp=f.readline().split()
    f.close()
    pdate=date(int(inp[0][6:]),int(inp[0][:2]),int(inp[0][3:5]))
    end_date=str(datetime.today()).split()[0]
    tdate1=date(int(end_date[:4]),int(end_date[5:7]),int(end_date[8:]))
    delta=timedelta(days=1)
    i=int(inp[1])
    while pdate<=tdate1:
        dat=pdate.strftime("%m-%d-%Y")
        try:
            data=setup_data(pd.read_csv(BASE_URL+dat+".csv"))
        except urllib.error.HTTPError:
            break
        wb['A'+str(i)]=dat
        for a in range(len(countries)):
            x=list(data.index)
            y=countries[a]
            if y in x:
                wb[l[a]+str(i)]=int(data.loc[y]["Active"])
            elif y not in x:
                f=False
                for b in range(len(x)):
                    if y in x[b]:
                        f=True
                        y=x[b]
                        break
                if f:
                    wb[l[a]+str(i)]=int(data.loc[y]["Active"])
                else:
                    wb[l[a]+str(i)]=0
            else:
                wb[l[a]+'2']=0
        pdate+=delta
        print(dat)
        i+=1
    g=open("STATUS.txt",'w')
    g.write(dat+' '+str(i))
    g.close()
    book.save(filename="data.xlsx")
